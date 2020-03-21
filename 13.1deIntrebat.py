import openpyxl, sys
x = int(input('Where you want to start:'))
y = int(input('How many blank rows:'))
wb = openpyxl.load_workbook('example.xlsx') 
sheet = wb.active
rows = tuple(sheet.rows)
sheet = wb['Sheet1']
for rowNum in rows[::-1]:
    for cellNum in rowNum:
        c = cellNum.column
        r = cellNum.row
        if r >= x and r < x+y:
            sheet.cell(row=r+y, column=c).value = sheet.cell(row=r, column=c).value
            sheet.cell(row=r, column=c).value = ''
        elif r >= x+y:
            sheet.cell(row=r+y, column=c).value = cellNum.value
wb.save('13dot1.xlsx')
sys.exit(0)