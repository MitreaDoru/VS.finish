import openpyxl, sys
x = int(input('Enter a number:'))
wb = openpyxl.Workbook() # Create a blank workbook.
wb.sheetnames
sheet = wb['Sheet']
for row1 in range(1,x):
    for column1 in range(1,x):
       sheet.cell(row=row1, column=column1).value = (row1)*(column1) 
wb.save('table_{0}.xlsx'.format(x))
sys.exit(0)

