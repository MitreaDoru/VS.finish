import openpyxl, sys, os
wb = openpyxl.load_workbook('example.xlsx')
sheet = wb.active
sheet = wb['Sheet1']
x = 0
y = 0
z = 0
for x in range(1,sheet.max_column+1):
    z += 1
    f = open('Coloana'+str(z)+'.txt', 'w+')
    for y in range (1,sheet.max_row+1):
        f.write(str(sheet.cell(row=y, column=x).value)+'\n')        
f.close()   
sys.exit(0)