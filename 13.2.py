import openpyxl, sys
wb = openpyxl.load_workbook('example.xlsx') 
sheet = wb.active
wb1 = openpyxl.Workbook() # Create a blank workbook.
wb1.sheetnames
sheet1 = wb1['Sheet']
sheet = wb['Sheet1']
for x in range(1,sheet.max_column+1):
    for y in range(1,sheet.max_row+1):
        sheet1.cell(row=x, column=y).value = sheet.cell(row=y, column=x).value 
wb1.save('13dot1.xlsx')
sys.exit(0)